# -*- coding: utf-8 -*-
"""
Created on 2024/9/23 15:56
@author: Whenxuan Wang
@email: wwhenxuan@gmail.com
用于模型训练的代码
"""
import os
from os import path
import sys
from time import sleep
from datetime import datetime

from numpy import round

import torch
from torch import nn
from torch import Tensor
from torch.optim import Optimizer
from torch.optim.lr_scheduler import LRScheduler

from torch_geometric.loader import DataLoader

import openpyxl as xl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

from matplotlib import pyplot as plt

from accelerate import Accelerator
from typing import Callable, Tuple
from colorama import Fore, Style
from tqdm import tqdm

from utils.tools import makedir
from utils.model_interface import ModelInterface
from utils.optimizer_interface import OptimInterface
from utils.data_interface import DataInterface
from utils.loss_fn import get_criterion


class Trainer(object):
    """模型训练的通用接口"""

    def __init__(
        self,
        args,
        model_interface: ModelInterface,
        optimizer_interface: OptimInterface,
        data_interface: DataInterface,
        criterion_interface: nn.Module,
        accelerator: Accelerator,
    ) -> None:
        self.args = args

        # 要进行预训练还是微调
        self.tasks = args.tasks

        # 获取训练轮数
        self.num_epochs = args.num_epochs
        # 预训练时保存模型的轮数
        self.save_epochs = args.save_epochs
        # 模型微调时开始验证的轮数
        self.test_epochs = args.test_epochs

        # 记录各种接口
        self.model_interface = model_interface
        self.optimizer_interface = optimizer_interface
        self.criterion_interface = criterion_interface

        # 获取协同加速器
        self.accelerator = accelerator
        # 记录当前的进程号
        self.process_index = self.accelerator.process_index

        # 获取训练集和验证集
        self.data_interface = data_interface
        # 获取当前训练设备
        self.device = self.accelerator.device

        # 记录Excel中的内容
        self.logging_info_finetuning = [
            "Time",
            "Model",
            "Test acc",
            "Test loss",
            "Train num",
            "Dataset",
            "Index",
            "Graph",
            "LR",
            "Batch size",
        ]
        self.logging_info_pretraining = ["Epoch", "Time", "Loss"]

        # 记录Excel中的各自样式
        black_side = Side(border_style="thin", color="000000")
        self.style = {
            "top_fill": PatternFill(start_color="DCE6F1", fill_type="solid"),
            "border": Border(
                left=black_side, right=black_side, top=black_side, bottom=black_side
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

        # 初始化用于logging的地址和目录
        self.main_path, self.params_path, self.wb_path = self.init_path()

    def init_path(self) -> Tuple[str, str, str]:
        """获取本次预训练保存模型和logging的地址"""
        # 保存模型的目录
        save_path = self.args.save_path

        if self.tasks == "pretrain":
            # 判断保存目录下有多少个文件
            num_folder = len(os.listdir(save_path))

            # 创建本次保存模型的文件夹
            folder_name = f"exp{num_folder}"
            makedir(save_path, folder_name)

            # 更新保存目录的主要地址
            main_path = path.join(save_path, folder_name)

            # 创建保存模型参数的文件夹
            makedir(main_path, "params")
            params_path = path.join(main_path, "params")
            print(f"Attention the logging path is {main_path}")

            # 创建记录预训练配置的Excel文件
            wb = self.init_excel()
            wb_path = path.join(main_path, "logging.xlsx")
            wb.save(wb_path)

        elif self.tasks == "finetune":
            # 创建记录模型微调的Excel
            files = os.listdir(save_path)
            # 检验Excel文件是否已经创建
            if "finetuning_results.xlsx" not in files:
                # 初始化表格
                wb = self.init_excel()
                wb.save(path.join(save_path, "finetuning_results.xlsx"))
            # 记录主要的地址信息
            main_path, params_path = None, None
            wb_path = path.join(save_path, "finetuning_results.xlsx")

        else:
            raise ValueError("Wrong task!")

        return main_path, params_path, wb_path

    def init_excel(self) -> xl.Workbook:
        """初始化记录训练过程的表格"""
        # 创建一个工作表
        wb = xl.Workbook()

        # 获取当前的工作表
        sheet = wb.active
        if self.tasks == "pretrain":
            # 如果执行的任务是预训练
            info_list = self.logging_info_pretraining
        elif self.tasks == "finetune":
            # 如果执行的任务是模型的微调
            info_list = self.logging_info_finetuning
        else:
            raise ValueError("Wrong task!")
        for column, info in enumerate(info_list, 1):
            sheet.cell(row=1, column=column).value = info
            sheet.cell(row=1, column=column).fill = self.style["top_fill"]
            sheet.cell(row=1, column=column).border = self.style["border"]
        return wb

    def run(self) -> None:
        """启动模型的预训练或微调"""
        if self.tasks == "pretrain":
            self.pretraining()
        elif self.tasks == "finetune":
            self.finetuning()
        else:
            raise ValueError("Wrong task!")

    def pretraining(self):
        """进行模型的预训练"""
        self.accelerator.print(
            Fore.GREEN + "Starting Model Pretraining..." + Style.RESET_ALL
        )
        # 记录本次预训练过程的损失
        train_loss = torch.zeros(self.num_epochs, device=self.device)

        # 获取模型
        model = self.model_interface.model
        # 获取模型可训练参数
        train_params = self.model_interface.trainable_params()

        # 获取预训练使用的数据集
        train_loader = self.data_interface.get_pretraining_dataloader()
        self.data_interface.pointer = 0

        # 获取损失函数
        criterion = self.criterion_interface

        # 获取优化器
        optimizer = self.optimizer_interface.load_optimizer(parameters=train_params)
        # 获取动态学习率调整
        scheduler = self.optimizer_interface.load_scheduler(
            optimizer, loader_len=len(train_loader)
        )

        # 开始模型的预训练
        for idx, epoch in enumerate(range(1, self.num_epochs + 1)):
            # 预训练一个Epoch
            train_loss[idx], model, optimizer, scheduler = self.pretrain_one_epoch(
                epoch=epoch,
                model=model,
                optimizer=optimizer,
                scheduler=scheduler,
                criterion=criterion,
            )

            # 判断是否需要保存模型
            if epoch % self.save_epochs == 0:
                self.save_model(epoch=epoch, model=model)

            # 登记本次模型的训练结果
            self.logging_pretraining(epoch=epoch, loss=float(train_loss[idx]))

        # 是否绘制预训练损失图像
        if self.args.plot_pretrain is True:
            self.plot_pretrain(train_loss=train_loss)

    def pretrain_one_epoch(
        self,
        epoch: int,
        model: nn.Module,
        optimizer: torch.optim.Optimizer,
        scheduler: torch.optim.lr_scheduler.LRScheduler,
        criterion: nn.Module,
    ) -> Tuple[
        float, nn.Module, torch.optim.Optimizer, torch.optim.lr_scheduler.LRScheduler
    ]:
        """预训练一个Epoch的函数"""
        # 用于计算累计损失
        loss_value = torch.zeros(1, device=self.device)
        num_samples = 0

        for ii in range(1, len(self.data_interface) + 1):
            # 在一个Epoch中要遍历读取完所有的数据
            self.accelerator.print(
                Fore.RED + "Now is loading pretraining data" + Style.RESET_ALL,
                end=" -> ",
            )

            # 获取最新的数据集
            train_loader = self.data_interface.get_pretraining_dataloader()
            # 将该数据集加入到多卡并行当中
            train_loader = self.accelerator.prepare_data_loader(
                train_loader, device_placement=True
            )
            sleep(2.56)

            self.accelerator.print(
                Fore.GREEN + "successfully loaded!" + Style.RESET_ALL
            )
            model.train()

            # 使用进度条对训练数据集对象进行封装
            data_loader = tqdm(train_loader, file=sys.stdout)

            # 开始遍历当前的数据集进行训练
            for step, (signals, graphs) in enumerate(data_loader, 1):
                # 清空优化器原有的梯度
                optimizer.zero_grad()
                num_samples += signals.size(0)

                # 通过模型的正向传播获得时间序列和图级别的特征
                time_features, graph_features, logit_scale = model(
                    x=signals, graph=graphs
                )
                # 通过CLIP的方式计算对比损失
                loss = criterion(graph_features, signals)

                # 误差的反向传播
                self.accelerator.backward(loss)
                # 参数的更新
                optimizer.step()
                # 检查模型损失
                check_loss(loss, train_type="Pretrain")

                # 记录本次的损失
                loss_value += loss.item()

                data_loader.desc = (
                    "["
                    + Fore.GREEN
                    + f"Epoch {epoch}"
                    + Style.RESET_ALL
                    + "] "
                    + f"Loss={loss_value.item() / num_samples:.5f}"
                )
                # 动态调整学习率
                scheduler.step()

        # 本轮预训练的损失值
        loss_value = self.accelerator.gather(loss_value).mean().item()
        loss_value = float(round(loss_value / num_samples, 4))

        return loss_value, model, optimizer, scheduler

    def logging_pretraining(self, epoch: int, loss: float) -> None:
        """用于记录模型的预训练过程"""
        # 登记本次模型训练的信息
        wb = load_workbook(self.wb_path)
        sheet = wb.active
        # 获取最后一行
        max_row = sheet.max_row + 1
        for col, info in enumerate(
            [
                epoch,
                str(datetime.now()),
                loss,
            ]
        ):
            # 记录本次微调的所有结果并改变样式
            sheet.cell(row=max_row, column=col).value = info
            sheet.cell(row=max_row, column=col).border = self.style["border"]

        # 保存本次logging的结果
        wb.save(self.wb_path)

    def save_model(self, epoch: int, model: nn.Module) -> None:
        """保存预训练模型的参数"""
        if self.process_index == 0:
            # 在主进程中保存模型参数
            self.accelerator.print(
                Fore.RED
                + "Now is saving the pre-trained model parameters"
                + Style.RESET_ALL,
                end=" -> ",
            )
            save_name = f"{epoch}.pth"
            torch.save(model.state_dict(), path.join(self.params_path, save_name))
            self.accelerator.print(Fore.GREEN + "successfully saved!" + Style.RESET_ALL)
        sleep(2.56)

    def plot_pretrain(self, train_loss) -> None:
        """绘制模型预训练损失图像"""
        fig, ax = plt.subplots(figsize=(12, 4))
        ax.plot(train_loss, color="royalblue")
        # 保存图像
        fig.savefig(
            path.join(self.main_path, "pretraining_losses.jpg"),
            dpi=800,
            bbox_inches="tight",
        )

    def finetuning(self):
        """进行模型的微调"""
        self.accelerator.print(
            Fore.GREEN + "Starting SymTime Model Fine-tuning..." + Style.RESET_ALL
        )
        # 记录微调训练过程中的结果
        train_loss, train_acc, test_loss, test_acc = (
            torch.zeros(self.num_epochs, device=self.device),
            torch.zeros(self.num_epochs, device=self.device),
            torch.zeros(self.num_epochs, device=self.device),
            torch.zeros(self.num_epochs, device=self.device),
        )

        # 获取模型
        model = self.model_interface.model
        # 获取模型可训练参数
        train_params = self.model_interface.trainable_params()

        # 获取训练和验证使用的数据集
        train_loader, test_loader = self.data_interface.get_finetuning_dataloader()

        # 获取损失函数
        criterion = self.criterion_interface

        # 获取优化器
        optimizer = self.optimizer_interface.load_optimizer(parameters=train_params)
        # 获取动态学习率调整
        scheduler = self.optimizer_interface.load_scheduler(
            optimizer, loader_len=len(train_loader)
        )

        # 为模型的训练进行准备
        (
            model,
            optimizer,
            scheduler,
            train_loader,
            test_loader,
        ) = self.accelerator.prepare(
            model, optimizer, scheduler, train_loader, test_loader
        )

        # 开始模型的训练
        for idx, epoch in enumerate(range(1, self.num_epochs + 1), 0):
            # 开始训练一个Epoch
            (
                train_loss[idx],
                train_acc[idx],
                model,
                optimizer,
                scheduler,
            ) = self.finetune_one_epoch(
                epoch=epoch,
                model=model,
                optimizer=optimizer,
                scheduler=scheduler,
                criterion=criterion,
                train_loader=train_loader,
            )
            # 开始进行模型的测试
            if epoch >= self.test_epochs:
                test_loss[idx], test_acc[idx] = self.evaluate(
                    epoch=epoch,
                    model=model,
                    test_loader=test_loader,
                    criterion=criterion,
                )

        # 结束模型的训练和验证，记录本次的训练信息
        index = torch.argmax(test_acc)
        accuracy = round(test_acc[index].item(), 5)
        loss = round(test_loss[index].item(), 5)
        self.logging_finetuning(accuracy=accuracy, loss=loss)

    def finetune_one_epoch(
        self,
        epoch: int,
        model: nn.Module,
        optimizer: torch.optim.Optimizer,
        scheduler: torch.optim.lr_scheduler.LRScheduler,
        criterion: Callable,
        train_loader: DataLoader,
    ) -> Tuple[
        float,
        float,
        nn.Module,
        torch.optim.Optimizer,
        torch.optim.lr_scheduler.LRScheduler,
    ]:
        """微调一个Epoch的函数"""
        num_samples = 0
        step = 1

        # 记录本次训练的结果
        accu_loss = torch.zeros(1, device=self.device)
        accu_num = torch.zeros(1, device=self.device)

        # 创建数据迭代器对象
        data_loader = tqdm(iterable=train_loader, file=sys.stdout)

        for step, (signals, graphs, labels) in enumerate(data_loader, 1):
            optimizer.zero_grad()
            # 记录本次的训练的样本数目
            num_samples += signals.size(0)

            # 进行模型的正向传播
            outputs = model(x=signals, graph=graphs)
            # 根据模型的输出计算损失
            # print(labels.size(), outputs.size())
            loss = criterion(outputs, labels.long())

            # 误差反向传播
            self.accelerator.backward(loss)
            # 通过优化器更新模型的参数
            optimizer.step()
            # 计算模型故障诊断的准确率
            _, predicted = torch.max(outputs, dim=1)

            # 记录本次训练的结果
            accu_num += torch.eq(predicted, labels).sum()
            accu_loss += loss.item()

            # 打印本次训练的情况
            data_loader.desc = f"[\033[32mTrain Epoch\033[0m {epoch}] loss: {round(accu_loss.item() / step, 4)}, acc: {round(accu_num.item() / num_samples, 4)}"
            # 检查损失是否正常
            check_loss(loss=loss, train_type="finetuning")

            # 更新学习率动态调整模块
            scheduler.step()

        return (
            accu_loss.item() / step,
            accu_num.item() / num_samples,
            model,
            optimizer,
            scheduler,
        )

    @torch.no_grad()
    def evaluate(
        self,
        epoch: int,
        model: nn.Module,
        criterion: Callable,
        test_loader: DataLoader,
    ) -> Tuple[float, float]:
        """在模型的微调时通过测试集进行评估"""
        model.eval()
        num_samples = 0

        # 记录本次验证的损失和准确率
        accu_loss = torch.zeros(1, device=self.device)
        accu_num = torch.zeros(1, device=self.device)

        # 对测试迭代器进行封装
        data_loader = tqdm(iterable=test_loader, file=sys.stdout)

        # 开始模型的验证
        with torch.no_grad():
            for step, (signals, graphs, labels) in enumerate(data_loader, 1):
                # 通过正向传播获得模型的输出
                outputs = model(x=signals, graph=graphs)
                num_samples += signals.size(0)

                # 计算验证损失
                loss = criterion(outputs, labels.long())
                # 计算验证准确率
                _, predicted = torch.max(outputs, dim=1)

                # 记录本次验证的结果
                accu_num += torch.eq(predicted, labels).sum()
                accu_loss += loss.item()

                # 打印本次训练的情况
                data_loader.desc = (
                    f"[\033[31mTest  Epoch\033[0m {epoch}] loss: {round(accu_loss.item() / step, 4)}, "
                    + Fore.GREEN
                    + f"acc: {round(accu_num.item() / num_samples, 4)}"
                    + Style.RESET_ALL
                )
                # 检查损失是否正常
                check_loss(loss=loss, train_type="finetuning")

        return accu_loss.item() / step, accu_num.item() / num_samples

    def logging_finetuning(self, accuracy: float, loss: float) -> None:
        """用于记录模型的微调过程"""
        # 登记本次模型训练的信息
        wb = load_workbook(self.wb_path)
        sheet = wb.active
        # 获取最后一行
        max_row = sheet.max_row + 1
        for col, info in enumerate(
            [
                str(datetime.now()),
                self.args.model,
                accuracy,
                loss,
                self.num_epochs,
                self.args.dataset,
                self.args.dataset_index,
                self.args.graph_generate,
                self.args.learning_rate,
                self.args.batch_size,
            ]
        ):
            # 记录本次微调的所有结果并改变样式
            sheet.cell(row=max_row, column=col).value = info
            sheet.cell(row=max_row, column=col).border = self.style["border"]

        # 保存本次logging的结果
        wb.save(self.wb_path)


class PreTrainer(object):
    """用于模型预训练的接口"""

    def __init__(
        self,
        args,
        model: nn.Module,
        optimizer: Optimizer,
        criterion: Callable,
        scheduler: LRScheduler,
        accelerator: Accelerator,
        data_interface: DataInterface,
    ):
        self.args = args
        # 获取训练轮数
        self.num_epochs = args.num_epochs
        # 预训练时保存模型的轮数
        self.save_epochs = args.save_epochs

        # TODO: 下面这些内容全部都要换 在类内进行prepare
        # 获取模型和训练数据集
        self.model = model

        # 获取神经网络的优化器
        self.optimizer = optimizer
        # 获取损失函数
        self.criterion = criterion
        # 获取动态调整学习率
        self.scheduler = scheduler

        # 获取协同加速器
        self.accelerator = accelerator
        # 记录当前的进程号
        self.process_index = self.accelerator.process_index

        # 获取训练集和验证集
        self.data_interface = data_interface
        # 获取当前训练设备
        self.device = self.accelerator.device

        # 初始化用于logging的地址和目录
        self.main_path, self.params_path, self.wb_path = self.init_path()

    def pretraining(self):
        pass

    def finetuning(self):
        """用于模型微调的接口文件"""
        self.accelerator.print(
            Fore.GREEN + "Starting SymTime Model Fine-tuning..." + Style.RESET_ALL
        )
        # 记录微调训练过程中的结果
        train_loss, train_acc, test_loss, test_acc = (
            torch.zeros(self.num_epochs, device=self.device),
            torch.zeros(self.num_epochs, device=self.device),
            torch.zeros(self.num_epochs, device=self.device),
            torch.zeros(self.num_epochs, device=self.device),
        )

        # 获取用于训练和验证的数据集
        train_loader, test_loader = self.data_interface.get_finetuning_dataloader()

        # 开始模型的微调
        for idx, epoch in enumerate(self.num_epochs):
            num_samples = 0  # 这一个Epoch中遍历的累计样本数目
            self.model.train()

    def finetuning_epoch(self, data_loader, epoch) -> Tuple:
        """微调一个epoch的函数"""
        # 设置模型为训练模型
        self.model.train()
        num_samples = 0  # 这一个Epoch中遍历的累计样本数目

        # 记录本次训练累计的损失
        accu_loss = torch.zeros(1, device=self.device)
        accu_num = torch.zeros(1, device=self.device)

        # 对微调使用的DataLoader对象进行封装
        data_loader = tqdm(data_loader, file=sys.stdout)

        # 开始本次训练
        for step, (features, labels, prompts) in enumerate(data_loader, 1):
            # 清空模型的梯度信息
            self.optimizer.zero_grad()
            features, labels = features.to(self.device), labels.to(self.device)
            outputs = self.model(features, prompts)
            num_samples += features.shape[0]
            # print("output", outputs.size(), outputs)
            # print("label", labels.size(), labels.float().long())
            loss = self.criterion(outputs, labels.long())
            self.accelerator.backward(loss)
            self.optimizer.step()
            _, predicted = torch.max(outputs, dim=1)
            # print(outputs.size(), predicted.size(), labels.size())
            accu_num += torch.eq(predicted, labels).sum()
            accu_loss += loss.item()
            data_loader.desc = f"[\033[32mTrain Epoch\033[0m {epoch}] loss: {round(accu_loss.item() / step, 4)}, acc: {round(accu_num.item() / num_samples, 4)}"
            self.__check_loss(loss=loss)
            self.scheduler.step()

        return accu_loss.item() / step, accu_num.item() / num_samples

    def fit(self) -> Tuple[Tensor, Tensor, Tensor, Tensor, Tensor]:
        """训练模型拟合数据"""
        self.accelerator.print(
            Fore.GREEN + "Starting SymTime Model Pretraining..." + Style.RESET_ALL
        )
        train_loss = torch.zeros(self.num_epochs, device=self.device)
        train_loss_mtm = torch.zeros(self.num_epochs, device=self.device)
        train_loss_mlm = torch.zeros(self.num_epochs, device=self.device)
        train_loss_t2s = torch.zeros(self.num_epochs, device=self.device)
        train_loss_s2t = torch.zeros(self.num_epochs, device=self.device)
        for idx, epoch in enumerate(range(1, self.num_epochs + 1), 0):
            """这里是开始了一个Epoch"""
            num_samples = 0  # 这一个Epoch中遍历的累计样本数目
            for ii in range(1, len(self.data_interface) + 1):
                """在一个Epoch中要遍历读取完所有的数据"""
                self.accelerator.print(
                    Fore.RED + "Now is loading pretraining data" + Style.RESET_ALL,
                    end=" -> ",
                )
                train_loader = self.data_interface.get_dataloader()
                train_loader = self.accelerator.prepare_data_loader(
                    train_loader, device_placement=True
                )
                sleep(2)
                self.accelerator.print(
                    Fore.GREEN + "successfully loaded!" + Style.RESET_ALL
                )
                self.model.train()
                data_loader = tqdm(train_loader, file=sys.stdout)
                for step, (time, time_mask, sym_ids, sym_mask) in enumerate(
                    data_loader, 1
                ):
                    self.optimizer.zero_grad()
                    num_samples += time.shape[0]
                    # 直接在模型正向传播的过程中获得损失
                    loss_mtm, loss_mlm, loss_t2s, loss_s2t = self.model(
                        time, time_mask, sym_ids, sym_mask
                    )
                    # 获取和整合误差
                    loss = loss_mtm + loss_mlm + (loss_t2s + loss_s2t) / 2
                    # 误差的反向传播
                    self.accelerator.backward(loss)
                    # 参数的更新
                    self.optimizer.step()
                    # 检查模型损失
                    check_loss(loss, train_type="Pretrain")
                    # 计算这个epoch的累计损失
                    train_loss[idx] += loss.item()
                    train_loss_mtm[idx] += loss_mtm.item()
                    train_loss_mlm[idx] += loss_mlm.item()
                    train_loss_t2s[idx] += loss_t2s.item()
                    train_loss_s2t[idx] += loss_s2t.item()
                    data_loader.desc = (
                        "["
                        + Fore.GREEN
                        + f"Epoch {epoch}"
                        + Style.RESET_ALL
                        + "] "
                        + "Loss="
                        + Fore.GREEN
                        + f"{round(train_loss[idx].item() / num_samples, 6)}"
                        + Style.RESET_ALL
                        + f" loss_mtm: {round(train_loss_mtm[idx].item() / num_samples, 6)}, loss_mlm: {round(train_loss_mlm[idx].item() / num_samples, 6)}, "
                        f"loss_t2s: {round(train_loss_t2s[idx].item() / num_samples, 6)}, loss_s2t: {round(train_loss_s2t[idx].item() / num_samples, 6)}"
                    )
                    # 动态调整学习率
                    self.scheduler.step()
                # 释放训练优化器的内存
                self.accelerator.clear(train_loader)
            # 记录最终损失的变化
            train_loss[idx] = train_loss[idx] / num_samples
            train_loss_mtm[idx] = train_loss_mtm[idx] / num_samples
            train_loss_mlm[idx] = train_loss_mlm[idx] / num_samples
            train_loss_t2s[idx] = train_loss_t2s[idx] / num_samples
            train_loss_s2t[idx] = train_loss_s2t[idx] / num_samples
            if epoch % self.save_epochs == 0:
                # 保存一次预训练模型的参数
                self.save_model(loss=train_loss[idx], epoch=epoch)
            # Logging训练过程 登记当前的epoch和最后的损失
            self.logging_epoch(
                epoch,
                train_loss[idx],
                train_loss_mtm[idx],
                train_loss_mlm[idx],
                train_loss_t2s[idx],
                train_loss_s2t[idx],
            )
        # """这部分可以调整一下专门写一个函数来执行"""
        # # 记录logging结果
        # self.logging.dict2csv()
        # self.logging.plot_results()
        return (
            train_loss,
            train_loss_mtm,
            train_loss_mlm,
            train_loss_t2s,
            train_loss_s2t,
        )

    def init_path(self) -> Tuple[str, str, str]:
        """获取本次预训练保存模型和logging的地址"""
        # 保存模型的目录
        save_path = self.args.save_path
        # 判断保存目录下有多少个文件
        num_folder = len(os.listdir(save_path))

        # 创建本次保存模型的文件夹
        folder_name = f"exp{num_folder + 1}"
        makedir(save_path, folder_name)

        # 更新保存目录的主要地址
        main_path = path.join(save_path, folder_name)

        # 创建保存模型参数的文件夹
        makedir(main_path, "params")
        params_path = path.join(main_path, "params")
        print(f"Attention the logging path is {main_path}")

        # 创建记录预训练配置的Excel文件
        wb = xl.Workbook()
        wb_path = path.join(main_path, "pretrain_logging.xlsx")
        wb.save(wb_path)

        return main_path, params_path, wb_path

    def save_model(self, epoch: int, loss: Tensor) -> None:
        """保存模型的参数"""
        if self.process_index == 0:
            self.accelerator.print(
                Fore.RED + "Now is saving the pretrained params" + Style.RESET_ALL,
                end=" -> ",
            )
            save_name = f"{epoch}_{round(loss.item(), 4)}.pth"
            torch.save(
                self.model.time_encoder.state_dict(),
                path.join(self.params_path, save_name),
            )
            self.accelerator.print(Fore.GREEN + "successfully saved!" + Style.RESET_ALL)

    def logging_epoch(self):
        pass


def check_loss(loss: Tensor, train_type: str) -> None:
    """检查训练和验证的损失避免梯度爆炸"""
    if not torch.isfinite(loss):
        print(
            Fore.RED + f"{train_type} now occurs ERROR: non-finite loss, end training!"
        )
        sys.exit(1)
